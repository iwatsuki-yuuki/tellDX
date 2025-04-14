import os
import uuid
from openai import OpenAI
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
import openpyxl
from linebot import LineBotApi
from linebot.models import TextSendMessage

load_dotenv()

app = FastAPI()

# CORS設定
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

# OpenAI設定
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# LINE API設定
LINE_CHANNEL_ACCESS_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
line_bot_api = LineBotApi(LINE_CHANNEL_ACCESS_TOKEN)

EXCEL_FILE = "data.xlsx"

# Excel初期設定（ファイルが存在しない場合）
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "担当者", "発注数", "文字起こし結果", "要約", "通知ステータス"])
    wb.save(EXCEL_FILE)

@app.post("/api/upload")
async def upload_audio(file: UploadFile = File(...)):
    file_id = str(uuid.uuid4())
    if not os.path.exists("uploaded_files"):
        os.mkdir("uploaded_files")
    file_path = os.path.join("uploaded_files", f"{file_id}.webm")

    # ファイル保存処理
    try:
        with open(file_path, "wb") as buffer:
            buffer.write(await file.read())
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"ファイル保存失敗: {str(e)}"})

    # Whisperで文字起こし
    try:
        with open(file_path, "rb") as audio_file:
            transcript = client.audio.transcriptions.create(
                model="whisper-1",
                file=audio_file
            )
        transcript_text = transcript.text
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": f"文字起こし失敗: {str(e)}"})

    # GPTで担当者と発注数を抽出
    try:
        extraction_prompt = (
            "以下の文章から担当者名と発注数を抽出してください。\n"
            "担当者名は以下の名前から必ず統一して選択：\n岩月, 伊藤, 佐藤, 鈴木, 田中\n"
            "フォーマット:「担当者:○○ 発注数:○個」\n\n"
            f"文章:\n{transcript_text}"
        )
        extraction_completion = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "担当者名と発注数を抽出する。"},
                {"role": "user", "content": extraction_prompt},
            ],
            max_tokens=50,
            temperature=0
        )
        extraction_result = extraction_completion.choices[0].message.content.strip()
        担当者 = extraction_result.split("担当者:")[1].split()[0].strip()
        発注数 = int(extraction_result.split("発注数:")[1].split("個")[0].strip())
    except Exception as e:
        担当者 = "不明"
        発注数 = 0

    # GPTで要約作成
    try:
        summary_completion = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "以下の文章を簡潔に要約してください。"},
                {"role": "user", "content": transcript_text},
            ],
            max_tokens=100,
            temperature=0.7
        )
        summary_text = summary_completion.choices[0].message.content.strip()
    except Exception as e:
        summary_text = "要約に失敗しました。"

    # Excel処理（担当者の検索とデータ書き込み）
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    line_user_id = None
    for row in ws.iter_rows(min_row=2):
        if row[1].value == 担当者:
            line_user_id = row[0].value
            break

    # LINE通知送信処理とステータス管理
    通知ステータス = "未通知"
    if line_user_id:
        try:
            message_text = f"担当者: {担当者}\n発注数: {発注数}個\n内容要約: {summary_text}"
            line_bot_api.push_message(line_user_id, TextSendMessage(text=message_text))
            通知ステータス = "通知済"
        except Exception as e:
            通知ステータス = "通知失敗"
            print(f"LINE通知失敗: {e}")

    # Excelへ明示的にデータを追加
    ws.append([
        line_user_id if line_user_id else "不明",  # ID
        担当者,                                 # 担当者
        発注数,                                 # 発注数
        transcript_text,                        # 文字起こし結果
        summary_text,                           # 要約
        通知ステータス                            # 通知ステータス
    ])
    wb.save(EXCEL_FILE)

    return {
        "file_id": file_id,
        "transcript": transcript_text,
        "summary": summary_text,
        "担当者": 担当者,
        "発注数": 発注数,
        "line_user_id": line_user_id,
        "通知ステータス": 通知ステータス
    }

@app.get("/api/data")
def get_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    data_list = [{header[i]: cell for i, cell in enumerate(row)} for row in data_rows]

    return {"data": data_list}